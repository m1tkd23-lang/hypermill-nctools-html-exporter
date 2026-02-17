# src\hypermill_nctools_html_exporter\export_xlsx.py
from __future__ import annotations

from dataclasses import asdict
from pathlib import Path
from typing import List, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

from .model import NcToolRecord


DEFAULT_COLUMNS = [
    # identity
    "nctool_no",
    "nctool_name",
    "nctool_comment",
    # assembly
    "holder_name",
    "holder_length",
    "tool_name",
    "tool_length",
    # tool page
    "tool_type",
    "tool_page_name",
    "tool_diameter_mm",
    "tool_corner_radius_mm",
    "tool_flutes",
    "tool_cut_length_ap_mm",
    "tool_shank_d_mm",
    "tool_chamfer_len_mm",
    "tool_tip_len_mm",
    "tool_taper_angle_deg",
    "spindle_rotation",
    # conditions
    "cond_S_n",
    "cond_FX",
    "cond_FZ",
    "cond_Fr",
    "cond_ap",
    "cond_ae",
    # image refs
    "image_rel_src",
    "image_abs_path",
    "image_cached_path",
    # provenance
    "source_html_path",
]


def _autosize_columns(ws, max_width: int = 60) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 10
        for r in range(1, min(ws.max_row, 200) + 1):
            v = ws.cell(row=r, column=col).value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[letter].width = min(max_width, max(10, max_len + 2))


def write_xlsx(
    records: List[NcToolRecord],
    out_xlsx: Path,
    embed_images: bool = True,
    image_col_name: str = "image_cached_path",
    image_cell_col: int | None = None,
    row_height: int = 90,
) -> Tuple[int, int]:
    """
    records -> XLSX
    Returns: (written_rows, embedded_images)
    """
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "nctools"

    cols = list(DEFAULT_COLUMNS)
    if embed_images and "image" not in cols:
        cols.append("image")

    ws.append(cols)

    img_count = 0
    for idx, rec in enumerate(records, start=2):
        d = asdict(rec)
        row = []
        for c in cols:
            if c == "image":
                row.append("")
            else:
                v = d.get(c, "")
                if hasattr(v, "__fspath__"):
                    v = str(v)
                row.append(v)
        ws.append(row)
        ws.row_dimensions[idx].height = row_height

    if embed_images:
        if image_cell_col is None:
            image_cell_col = cols.index("image") + 1

        for r, rec in enumerate(records, start=2):
            p = rec.image_cached_path
            if not p:
                continue
            try:
                img = XLImage(str(p))
                cell = ws.cell(row=r, column=image_cell_col)
                ws.add_image(img, cell.coordinate)
                img_count += 1
            except Exception:
                pass

        ws.column_dimensions[get_column_letter(image_cell_col)].width = 18

    _autosize_columns(ws)

    ws_meta = wb.create_sheet("meta")
    ws_meta.append(["records", len(records)])
    ws_meta.append(["embedded_images", img_count])
    ws_meta.append(["embed_images", str(embed_images)])

    ws_err = wb.create_sheet("errors")
    ws_err.append(["row_index(1-based in nctools)", "nctool_name", "message"])

    wb.save(out_xlsx)
    return len(records), img_count
