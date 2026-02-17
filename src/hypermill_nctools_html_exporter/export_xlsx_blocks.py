from __future__ import annotations

from pathlib import Path
from typing import List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.units import pixels_to_EMU, points_to_pixels

from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D

from .model import NcToolRecord


def _fit_columns(ws, widths: dict[int, float]) -> None:
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def _safe_str(x) -> str:
    return "" if x is None else str(x)


def _col_width_to_pixels(width: float) -> int:
    if width <= 0:
        return 0
    return int(width * 7.5 + 5)


def _row_height_to_pixels(points: float) -> int:
    return int(points_to_pixels(points))


def _apply_block_border(ws, start_row: int, end_row: int, max_col: int, img_col: int) -> None:
    thin = Side(style="thin")
    thick = Side(style="medium")
    none = Side(style=None)

    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            left = thin
            right = thin
            top = thin
            bottom = thin

            if r == start_row:
                top = thick
            if r == end_row:
                bottom = thick
            if c == 1:
                left = thick
            if c == max_col:
                right = thick

            # 画像列は縦結合なので内部横線を消す
            if c == img_col:
                if r != start_row:
                    top = none
                if r != end_row:
                    bottom = none

            ws.cell(r, c).border = Border(left=left, right=right, top=top, bottom=bottom)


def _center_image_in_merged_cell(ws, img: XLImage, col: int, row_top: int, row_bottom: int) -> None:
    col_letter = get_column_letter(col)

    col_width = ws.column_dimensions[col_letter].width or 8.43
    cell_w_px = _col_width_to_pixels(col_width)

    total_h_px = 0
    for r in range(row_top, row_bottom + 1):
        h_pt = ws.row_dimensions[r].height or 15
        total_h_px += _row_height_to_pixels(h_pt)

    img_w_px = int(getattr(img, "width", 0) or 0)
    img_h_px = int(getattr(img, "height", 0) or 0)

    off_x_px = max(0, (cell_w_px - img_w_px) // 2)
    off_y_px = max(0, (total_h_px - img_h_px) // 2)

    # 微調整（環境差吸収）
    off_x_px = max(0, off_x_px + 12)

    marker = AnchorMarker(
        col=col - 1,
        colOff=pixels_to_EMU(off_x_px),
        row=row_top - 1,
        rowOff=pixels_to_EMU(off_y_px),
    )
    size = XDRPositiveSize2D(
        cx=pixels_to_EMU(img_w_px),
        cy=pixels_to_EMU(img_h_px),
    )
    img.anchor = OneCellAnchor(_from=marker, ext=size)


def export_blocks_f2_xlsx(
    records: List[NcToolRecord],
    out_xlsx: Path,
    *,
    embed_images: bool = True,
    block_rows: int = 3,
    start_row: int = 2,
) -> Tuple[int, int]:
    """
    ヘッダー:
      No / NCツール名 / 呼径 / 識別 / 補正H / 補正D / 画像 / 種別 / 名称 / 詳細 / 追記

    呼径/識別/補正H/補正D は手入力欄なので常に空で出力する。
    """
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    # columns
    COL_NO = 1
    COL_NCNAME = 2
    COL_CALIBER = 3   # 呼径（手入力）
    COL_IDENT = 4     # 識別（手入力）
    COL_H = 5         # 補正H（手入力）
    COL_D = 6         # 補正D（手入力）
    COL_IMG = 7       # 画像
    COL_KIND = 8      # 種別
    COL_COMP = 9      # 名称
    COL_DETAIL = 10   # 詳細
    COL_NOTE = 11     # 追記（手入力）

    sep = " / "

    _fit_columns(
        ws,
        {
            COL_NO: 6,
            COL_NCNAME: 24,
            COL_CALIBER: 7,  # ★指定
            COL_IDENT: 7,    # ★指定
            COL_H: 7,        # ★指定
            COL_D: 7,        # ★指定
            COL_IMG: 40,
            COL_KIND: 12,
            COL_COMP: 55,
            COL_DETAIL: 55,
            COL_NOTE: 50,    # ★指定
        },
    )

    font_header = Font(size=11, bold=True)
    font_title = Font(size=14, bold=True)
    font_norm = Font(size=11)
    align_vcenter_left = Alignment(vertical="center", horizontal="left", wrap_text=True)

    # header
    ws.row_dimensions[1].height = 22
    headers = [
        "No", "NCツール名", "呼径", "識別", "補正H", "補正D",
        "画像", "種別", "名称", "詳細", "追記"
    ]
    for i, text in enumerate(headers, start=1):
        cell = ws.cell(1, i)
        cell.value = text
        cell.font = font_header
        cell.alignment = align_vcenter_left
    _apply_block_border(ws, 1, 1, COL_NOTE, img_col=COL_IMG)

    img_count = 0
    written = 0

    for rec in records:
        r1 = start_row + written * block_rows
        r2, r3 = r1 + 1, r1 + 2

        for rr in (r1, r2, r3):
            ws.row_dimensions[rr].height = 80

        # merge (tool-wide columns)
        for c in (COL_NO, COL_NCNAME, COL_CALIBER, COL_IDENT, COL_H, COL_D, COL_IMG, COL_NOTE):
            ws.merge_cells(start_row=r1, start_column=c, end_row=r3, end_column=c)

        # common fields (write only at r1)
        ws.cell(r1, COL_NO).value = rec.nctool_no
        ws.cell(r1, COL_NO).font = Font(size=12, bold=True)

        ws.cell(r1, COL_NCNAME).value = rec.nctool_name
        ws.cell(r1, COL_NCNAME).font = font_title

        # 手入力欄：空で固定
        ws.cell(r1, COL_CALIBER).value = ""
        ws.cell(r1, COL_IDENT).value = ""
        ws.cell(r1, COL_H).value = ""
        ws.cell(r1, COL_D).value = ""
        ws.cell(r1, COL_NOTE).value = ""

        # row1 holder
        holder = rec.holder_name or rec.holder_page_name
        ws.cell(r1, COL_KIND).value = "holder"
        ws.cell(r1, COL_COMP).value = _safe_str(holder)
        ws.cell(r1, COL_DETAIL).value = (
            f"直径: {_safe_str(rec.tool_diameter_mm)}{sep}"
            f"刃数: {_safe_str(rec.tool_flutes)}{sep}"
            f"R: {_safe_str(rec.tool_corner_radius_mm)}\n"
            f"シャンク: {_safe_str(rec.tool_shank_d_mm)}\n"
            f"テーパー角: {_safe_str(rec.tool_taper_angle_deg)}{sep}"
            f"回転: {_safe_str(rec.spindle_rotation)}"
        )

        # row2 extension
        ws.cell(r2, COL_KIND).value = "extension"
        ws.cell(r2, COL_COMP).value = _safe_str(getattr(rec, "extensions_str", ""))
        ws.cell(r2, COL_DETAIL).value = (
            f"extension突き出し: {_safe_str(getattr(rec, 'ext_overhang_mm', '0'))}{sep}"
            f"工具突き出し: {_safe_str(getattr(rec, 'tool_overhang_mm', ''))}"
        )

        # row3 tool
        tool = rec.tool_page_name or rec.tool_name
        ws.cell(r3, COL_KIND).value = "tool"
        ws.cell(r3, COL_COMP).value = _safe_str(tool)
        ws.cell(r3, COL_DETAIL).value = f"突き出し長さ: {_safe_str(getattr(rec, 'overhang_mm', ''))}"

        # style
        for rr in (r1, r2, r3):
            for cc in range(COL_NO, COL_NOTE + 1):
                cell = ws.cell(rr, cc)
                if cell.font is None or cell.font == Font():
                    cell.font = font_norm
                cell.alignment = align_vcenter_left

        # image
        if embed_images and rec.image_cached_path:
            try:
                img = XLImage(str(rec.image_cached_path))
                ws.add_image(img, ws.cell(r1, COL_IMG).coordinate)
                _center_image_in_merged_cell(ws, img, col=COL_IMG, row_top=r1, row_bottom=r3)
                img_count += 1
            except Exception:
                pass

        _apply_block_border(ws, r1, r3, COL_NOTE, img_col=COL_IMG)
        written += 1

    wb.save(out_xlsx)
    return written, img_count
